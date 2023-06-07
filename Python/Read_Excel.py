import openpyxl, os

workbook = openpyxl.load_workbook('example.xlsx') #you can also give full path if file is not in cwd
print(type(workbook))

allsheet = workbook.sheetnames
print(type(allsheet))
print(allsheet)

for name in allsheet:
    print(name)


print(workbook.worksheets[0])
sheet = workbook.worksheets[0]
print(str(sheet['A2'].value))

print(sheet.cell(row=3, column=2).value)

#-----------------------editing----------------------------------------

workbook.create_sheet('Bobby')
sheet = workbook.worksheets[4]
sheet['A1'] = 'hello'
sheet['B1'] = 34
sheet.title = 'new bobby' #rename the sheet from bobby to new bobby
workbook.save('example.xlsx')